from typing import Union, Any
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class FinancialApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Gestión Financiera - Excel con Tkinter")
        self.root.geometry("800x600")
        self.root.configure(bg='#f0f0f0')

        # Excel
        self.wb = openpyxl.Workbook()
        self.worksheet = self.wb.active

        # Listas que guardan información ingresada por usuario
        self.gastos = []
        self.descrs = []
        self.ventas = []
        self.prodts = []
        self.names = []
        self.charges = []
        self.prod_chrgs = []
        self.name_chrgs = []

        self.fill_pattern = PatternFill(patternType='solid', fgColor='C64747')

        # Variables para totales
        self.total_gastos = 0
        self.total_ventas = 0
        self.total_encargos = 0

        self.setup_ui()

    def setup_ui(self):
        # Título principal
        title_label = tk.Label(self.root, text="GESTIÓN FINANCIERA",
                               font=("Arial", 20, "bold"),
                               bg='#f0f0f0', fg='#333')
        title_label.pack(pady=20)

        # Frame principal con pestañas
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=20, pady=10)

        # Pestañas
        self.create_gastos_tab()
        self.create_ventas_tab()
        self.create_encargos_tab()
        self.create_balance_tab()

        # Botón para guardar Excel
        save_button = tk.Button(self.root, text="Guardar Excel",
                                command=self.save_excel,
                                bg='#4CAF50', fg='white',
                                font=("Arial", 12, "bold"),
                                padx=20, pady=10)
        save_button.pack(pady=10)

    def create_gastos_tab(self):
        # Pestaña de Gastos
        gastos_frame = ttk.Frame(self.notebook)
        self.notebook.add(gastos_frame, text="Gastos")

        # Frame para ingresar gastos
        input_frame = tk.Frame(gastos_frame, bg='#fff', relief='raised', bd=2)
        input_frame.pack(fill='x', padx=10, pady=10)

        tk.Label(input_frame, text="INGRESAR GASTO",
                 font=("Arial", 14, "bold"), bg='#fff').pack(pady=10)

        # Campos de entrada
        tk.Label(input_frame, text="Gasto ($):", bg='#fff').pack()
        self.gasto_entry = tk.Entry(input_frame, width=30)
        self.gasto_entry.pack(pady=5)

        tk.Label(input_frame, text="Descripción:", bg='#fff').pack()
        self.descr_entry = tk.Entry(input_frame, width=30)
        self.descr_entry.pack(pady=5)

        tk.Button(input_frame, text="Agregar Gasto",
                  command=self.add_gasto,
                  bg='#FF6B6B', fg='white',
                  font=("Arial", 10, "bold")).pack(pady=10)

        # Frame para mostrar gastos
        display_frame = tk.Frame(gastos_frame, bg='#fff', relief='raised', bd=2)
        display_frame.pack(fill='both', expand=True, padx=10, pady=10)

        tk.Label(display_frame, text="LISTA DE GASTOS",
                 font=("Arial", 14, "bold"), bg='#fff').pack(pady=10)

        self.gastos_text = scrolledtext.ScrolledText(display_frame, height=10, width=70)
        self.gastos_text.pack(fill='both', expand=True, padx=10, pady=10)

        # Label para total
        self.total_gastos_label = tk.Label(display_frame, text="Total Gastos: $0.00",
                                           font=("Arial", 12, "bold"),
                                           bg='#fff', fg='#FF6B6B')
        self.total_gastos_label.pack(pady=10)

    def create_ventas_tab(self):
        # Pestaña de Ventas
        ventas_frame = ttk.Frame(self.notebook)
        self.notebook.add(ventas_frame, text="Ventas")

        # Frame para ingresar ventas
        input_frame = tk.Frame(ventas_frame, bg='#fff', relief='raised', bd=2)
        input_frame.pack(fill='x', padx=10, pady=10)

        tk.Label(input_frame, text="INGRESAR VENTA",
                 font=("Arial", 14, "bold"), bg='#fff').pack(pady=10)

        # Campos de entrada
        tk.Label(input_frame, text="Venta ($):", bg='#fff').pack()
        self.venta_entry = tk.Entry(input_frame, width=30)
        self.venta_entry.pack(pady=5)

        tk.Label(input_frame, text="Producto:", bg='#fff').pack()
        self.prod_entry = tk.Entry(input_frame, width=30)
        self.prod_entry.pack(pady=5)

        tk.Label(input_frame, text="Nombre del Comprador:", bg='#fff').pack()
        self.name_entry = tk.Entry(input_frame, width=30)
        self.name_entry.pack(pady=5)

        tk.Button(input_frame, text="Agregar Venta",
                  command=self.add_venta,
                  bg='#4CAF50', fg='white',
                  font=("Arial", 10, "bold")).pack(pady=10)

        # Frame para mostrar ventas
        display_frame = tk.Frame(ventas_frame, bg='#fff', relief='raised', bd=2)
        display_frame.pack(fill='both', expand=True, padx=10, pady=10)

        tk.Label(display_frame, text="LISTA DE VENTAS",
                 font=("Arial", 14, "bold"), bg='#fff').pack(pady=10)

        self.ventas_text = scrolledtext.ScrolledText(display_frame, height=10, width=70)
        self.ventas_text.pack(fill='both', expand=True, padx=10, pady=10)

        # Label para total
        self.total_ventas_label = tk.Label(display_frame, text="Total Ventas: $0.00",
                                           font=("Arial", 12, "bold"),
                                           bg='#fff', fg='#4CAF50')
        self.total_ventas_label.pack(pady=10)

    def create_encargos_tab(self):
        # Pestaña de Encargos
        encargos_frame = ttk.Frame(self.notebook)
        self.notebook.add(encargos_frame, text="Encargos")

        # Frame para ingresar encargos
        input_frame = tk.Frame(encargos_frame, bg='#fff', relief='raised', bd=2)
        input_frame.pack(fill='x', padx=10, pady=10)

        tk.Label(input_frame, text="INGRESAR ENCARGO",
                 font=("Arial", 14, "bold"), bg='#fff').pack(pady=10)

        # Campos de entrada
        tk.Label(input_frame, text="Valor del Encargo ($):", bg='#fff').pack()
        self.charge_entry = tk.Entry(input_frame, width=30)
        self.charge_entry.pack(pady=5)

        tk.Label(input_frame, text="Producto Encargado:", bg='#fff').pack()
        self.prod_chrg_entry = tk.Entry(input_frame, width=30)
        self.prod_chrg_entry.pack(pady=5)

        tk.Label(input_frame, text="Nombre del Comprador:", bg='#fff').pack()
        self.name_chrg_entry = tk.Entry(input_frame, width=30)
        self.name_chrg_entry.pack(pady=5)

        tk.Button(input_frame, text="Agregar Encargo",
                  command=self.add_encargo,
                  bg='#9C27B0', fg='white',
                  font=("Arial", 10, "bold")).pack(pady=10)

        # Frame para mostrar encargos
        display_frame = tk.Frame(encargos_frame, bg='#fff', relief='raised', bd=2)
        display_frame.pack(fill='both', expand=True, padx=10, pady=10)

        tk.Label(display_frame, text="LISTA DE ENCARGOS",
                 font=("Arial", 14, "bold"), bg='#fff').pack(pady=10)

        self.encargos_text = scrolledtext.ScrolledText(display_frame, height=10, width=70)
        self.encargos_text.pack(fill='both', expand=True, padx=10, pady=10)

        # Label para total
        self.total_encargos_label = tk.Label(display_frame, text="Total Encargos: $0.00",
                                             font=("Arial", 12, "bold"),
                                             bg='#fff', fg='#9C27B0')
        self.total_encargos_label.pack(pady=10)

    def create_balance_tab(self):
        # Pestaña de Balance
        balance_frame = ttk.Frame(self.notebook)
        self.notebook.add(balance_frame, text="Balance")

        # Frame para mostrar balance
        display_frame = tk.Frame(balance_frame, bg='#fff', relief='raised', bd=2)
        display_frame.pack(fill='both', expand=True, padx=10, pady=10)

        tk.Label(display_frame, text="RESUMEN FINANCIERO",
                 font=("Arial", 16, "bold"), bg='#fff').pack(pady=20)

        # Labels para mostrar totales
        self.balance_gastos_label = tk.Label(display_frame, text="Total Gastos: $0.00",
                                             font=("Arial", 14), bg='#fff', fg='#FF6B6B')
        self.balance_gastos_label.pack(pady=10)

        self.balance_ventas_label = tk.Label(display_frame, text="Total Ventas: $0.00",
                                             font=("Arial", 14), bg='#fff', fg='#4CAF50')
        self.balance_ventas_label.pack(pady=10)

        self.balance_encargos_label = tk.Label(display_frame, text="Total Encargos: $0.00",
                                               font=("Arial", 14), bg='#fff', fg='#9C27B0')
        self.balance_encargos_label.pack(pady=10)

        # Línea separadora
        separator = tk.Frame(display_frame, height=2, bg='#ccc')
        separator.pack(fill='x', padx=50, pady=20)

        # Balance final
        self.balance_final_label = tk.Label(display_frame, text="Balance (Ventas - Gastos): $0.00",
                                            font=("Arial", 16, "bold"), bg='#fff', fg='#2196F3')
        self.balance_final_label.pack(pady=10)

        self.balance_total_label = tk.Label(display_frame, text="Balance + Encargos: $0.00",
                                            font=("Arial", 16, "bold"), bg='#fff', fg='#FF9800')
        self.balance_total_label.pack(pady=10)

        # Botón para actualizar balance
        tk.Button(display_frame, text="Actualizar Balance",
                  command=self.update_balance,
                  bg='#2196F3', fg='white',
                  font=("Arial", 12, "bold")).pack(pady=20)

    # Métodos auxiliares (mantenidos del código original)
    def list_sumation(self, a):
        if not a:
            return 0
        else:
            return a[0] + self.list_sumation(a[1:])

    def to_excel(self, data, c):
        for i in range(len(data)):
            self.worksheet.cell(row=i + 1, column=c, value=data[i])

    def to_excel_op(self, head, row1, data, row2, c):
        self.worksheet.cell(row=row1, column=c, value=head)
        self.worksheet.cell(row=row2, column=c, value=data)

    def header(self, lista, valor):
        if valor in lista:
            lista.remove(valor)
        lista.insert(0, valor)
        return lista

    def background_color(self, worksheet, start_row, end_row, c1, c2, color):
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        if isinstance(end_row, list):
            end = len(end_row) + 1
        else:
            end = end_row

        for row in range(start_row, end):
            for col in range(c1, c2):
                col_letter = get_column_letter(col)
                cell = worksheet[f'{col_letter}{row}']
                cell.fill = fill

    def adjust_column_width(self, worksheet):
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 1

    # Métodos para agregar datos
    def add_gasto(self):
        try:
            gasto = float(self.gasto_entry.get())
            descr = self.descr_entry.get()

            if not descr:
                messagebox.showerror("Error", "Por favor ingrese una descripción")
                return

            self.gastos.insert(0, gasto)
            self.header(self.gastos, "Gastos")
            self.descrs.insert(0, descr)
            self.header(self.descrs, "Descripcion")

            self.to_excel(self.gastos, 1)
            self.to_excel(self.descrs, 2)
            self.background_color(self.worksheet, 2, self.gastos, 1, 3, 'E9C71B')
            self.background_color(self.worksheet, 1, 2, 1, 3, 'E9921B')

            self.update_gastos_display()
            self.clear_gastos_entries()

            messagebox.showinfo("Éxito", "Gasto agregado correctamente")

        except ValueError:
            messagebox.showerror("Error", "Por favor ingrese un valor numérico válido")

    def add_venta(self):
        try:
            venta = float(self.venta_entry.get())
            prod = self.prod_entry.get()
            name = self.name_entry.get()

            if not prod or not name:
                messagebox.showerror("Error", "Por favor complete todos los campos")
                return

            self.ventas.insert(0, venta)
            self.header(self.ventas, "Ventas")
            self.prodts.insert(0, prod)
            self.header(self.prodts, "Productos")
            self.names.insert(0, name)
            self.header(self.names, "Nombre del comprador")

            self.to_excel(self.ventas, 4)
            self.to_excel(self.prodts, 5)
            self.to_excel(self.names, 6)
            self.background_color(self.worksheet, 2, self.ventas, 4, 7, '58DB4B')
            self.background_color(self.worksheet, 1, 2, 4, 7, '1CBD0C')

            self.update_ventas_display()
            self.clear_ventas_entries()

            messagebox.showinfo("Éxito", "Venta agregada correctamente")

        except ValueError:
            messagebox.showerror("Error", "Por favor ingrese un valor numérico válido")

    def add_encargo(self):
        try:
            charge = float(self.charge_entry.get())
            prod_chrg = self.prod_chrg_entry.get()
            name_chrg = self.name_chrg_entry.get()

            if not prod_chrg or not name_chrg:
                messagebox.showerror("Error", "Por favor complete todos los campos")
                return

            self.charges.insert(0, charge)
            self.header(self.charges, "Valor del encargo")
            self.prod_chrgs.insert(0, prod_chrg)
            self.header(self.prod_chrgs, "Producto encargado")
            self.name_chrgs.insert(0, name_chrg)
            self.header(self.name_chrgs, "Nombre del comprador")

            self.to_excel(self.charges, 9)
            self.to_excel(self.prod_chrgs, 10)
            self.to_excel(self.name_chrgs, 11)
            self.background_color(self.worksheet, 2, self.charges, 9, 12, 'BB09F9')
            self.background_color(self.worksheet, 1, 2, 9, 12, '9805CF')

            self.update_encargos_display()
            self.clear_encargos_entries()

            messagebox.showinfo("Éxito", "Encargo agregado correctamente")

        except ValueError:
            messagebox.showerror("Error", "Por favor ingrese un valor numérico válido")

    # Métodos para limpiar campos
    def clear_gastos_entries(self):
        self.gasto_entry.delete(0, tk.END)
        self.descr_entry.delete(0, tk.END)

    def clear_ventas_entries(self):
        self.venta_entry.delete(0, tk.END)
        self.prod_entry.delete(0, tk.END)
        self.name_entry.delete(0, tk.END)

    def clear_encargos_entries(self):
        self.charge_entry.delete(0, tk.END)
        self.prod_chrg_entry.delete(0, tk.END)
        self.name_chrg_entry.delete(0, tk.END)

    # Métodos para actualizar displays
    def update_gastos_display(self):
        gastos_temp = [x for x in self.gastos if x != "Gastos"]
        descrs_temp = [x for x in self.descrs if x != "Descripcion"]

        self.gastos_text.delete(1.0, tk.END)
        for i, (gasto, descr) in enumerate(zip(gastos_temp, descrs_temp)):
            self.gastos_text.insert(tk.END, f"{i + 1}. ${gasto:.2f} - {descr}\n")

        self.total_gastos = self.list_sumation(gastos_temp)
        self.total_gastos_label.config(text=f"Total Gastos: ${self.total_gastos:.2f}")

    def update_ventas_display(self):
        ventas_temp = [x for x in self.ventas if x != "Ventas"]
        prodts_temp = [x for x in self.prodts if x != "Productos"]
        names_temp = [x for x in self.names if x != "Nombre del comprador"]

        self.ventas_text.delete(1.0, tk.END)
        for i, (venta, prod, name) in enumerate(zip(ventas_temp, prodts_temp, names_temp)):
            self.ventas_text.insert(tk.END, f"{i + 1}. ${venta:.2f} - {prod} - {name}\n")

        self.total_ventas = self.list_sumation(ventas_temp)
        self.total_ventas_label.config(text=f"Total Ventas: ${self.total_ventas:.2f}")

    def update_encargos_display(self):
        charges_temp = [x for x in self.charges if x != "Valor del encargo"]
        prod_chrgs_temp = [x for x in self.prod_chrgs if x != "Producto encargado"]
        name_chrgs_temp = [x for x in self.name_chrgs if x != "Nombre del comprador"]

        self.encargos_text.delete(1.0, tk.END)
        for i, (charge, prod, name) in enumerate(zip(charges_temp, prod_chrgs_temp, name_chrgs_temp)):
            self.encargos_text.insert(tk.END, f"{i + 1}. ${charge:.2f} - {prod} - {name}\n")

        self.total_encargos = self.list_sumation(charges_temp)
        self.total_encargos_label.config(text=f"Total Encargos: ${self.total_encargos:.2f}")

    def update_balance(self):
        # Actualizar labels de balance
        self.balance_gastos_label.config(text=f"Total Gastos: ${self.total_gastos:.2f}")
        self.balance_ventas_label.config(text=f"Total Ventas: ${self.total_ventas:.2f}")
        self.balance_encargos_label.config(text=f"Total Encargos: ${self.total_encargos:.2f}")

        balance = self.total_ventas - self.total_gastos
        self.balance_final_label.config(text=f"Balance (Ventas - Gastos): ${balance:.2f}")

        balance_total = balance + self.total_encargos
        self.balance_total_label.config(text=f"Balance + Encargos: ${balance_total:.2f}")

        # Actualizar Excel con totales
        if self.gastos:
            cel = 2
            cant = len([x for x in self.gastos if x != "Gastos"]) + 1
            self.to_excel_op("Gastos Totales", 1, f'=SUM(A{cel}:A{cant})', 2, 13)
            self.background_color(self.worksheet, 1, 3, 13, 14, 'E9921B')

        if self.ventas:
            cel = 2
            cant = len([x for x in self.ventas if x != "Ventas"]) + 1
            self.to_excel_op("Ventas Totales", 1, f'=SUM(D{cel}:D{cant})', 2, 14)
            self.background_color(self.worksheet, 1, 3, 14, 15, '1CBD0C')

        if self.charges:
            cel = 2
            cant = len([x for x in self.charges if x != "Valor del encargo"]) + 1
            self.to_excel_op("Total de encargos", 1, f'=SUM(I{cel}:I{cant})', 2, 15)
            self.background_color(self.worksheet, 1, 3, 15, 16, '9805CF')

        if self.ventas and self.gastos:
            self.to_excel_op("Balance", 5, "=N2-M2", 6, 13)
            self.background_color(self.worksheet, 5, 7, 13, 14, '08CCDF')

        if self.ventas and self.gastos and self.charges:
            self.to_excel_op("Balance mas total de encargos", 1, "=O2+M6", 2, 16)
            self.background_color(self.worksheet, 1, 3, 16, 17, '0891DF')

    def save_excel(self):
        try:
            self.adjust_column_width(self.worksheet)
            self.wb.save('fpdsvasos.xlsx')
            messagebox.showinfo("Éxito", "Archivo Excel guardado como 'fpdsvasos.xlsx'")
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar el archivo: {str(e)}")

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = FinancialApp()
    app.run()